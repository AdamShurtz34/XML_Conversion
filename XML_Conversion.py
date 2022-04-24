from openpyxl import load_workbook
from yattag import Doc, indent
from datetime import *
import dateutil.relativedelta as REL


space = "    "
def main():
    # Load our Excel File
    fileName = input("Enter file name (do not include .xlsx): ")
    fileName = fileName + ".xlsx"
    wb = load_workbook(fileName)
    # Getting an object of active sheet 1
    ws = wb.worksheets[0]

    today = datetime.now()
    rd = REL.relativedelta(days=1, weekday=REL.FR)
    next_friday = today + rd
    numOfTransactions = 0
    amtSum = 0
    for row in ws.iter_rows(min_row=2, max_col=13):
        numOfTransactions += 1
        amtSum = amtSum + row[11].value
    amtSum2 = "{:.2f}".format(amtSum)

    PaymentID1 = fileName[31:]
    PaymentID2 = PaymentID1[:-5]
    PaymentIDLyst = PaymentID2.split('-')
    PaymentID = ''
    for i in PaymentIDLyst:
        PaymentID = PaymentID + i

    f = open("XML "+next_friday.strftime("%#m-%#d-%Y")+".xml", "a")

    #TOP LEVEL ***********
    f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    f.write('<Document xmlns="urn:iso:std:iso:20022:tech:xsd:pain.001.001.03" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n')
    #level 1
    f.write(space+'<CstmrCdtTrfInitn>\n')
    

    #GROUP HEADER ****************
    #level 2
    f.write(space*2+'<GrpHdr>\n')
    #level 3
    f.write(space*3+'<MsgId>Commissions '+next_friday.strftime("%Y-%m-%d")+'</MsgId>\n')
    f.write(space*3+'<CreDtTm>'+today.strftime("%Y-%m-%dT%H:%M:%S")+'</CreDtTm>\n')
    f.write(space*3+'<NbOfTxs>'+str(numOfTransactions)+'</NbOfTxs>\n')
    f.write(space*3+'<CtrlSum>'+amtSum2+'</CtrlSum>\n')
    f.write(space*3+'<InitgPty>\n')
    #level 4
    f.write(space*4+'<Nm>Bank Name</Nm>\n')
    #end level 3
    f.write(space*3+'</InitgPty>\n')
    #END GROUP HEADER
    #end level 2
    f.write(space*2+'</GrpHdr>\n')
    

    #PAYMENT INFO ****************
    #level 2
    f.write(space*2+'<PmtInf>\n')
    #level 3
    f.write(space*3+'<PmtInfId>'+PaymentID+'</PmtInfId>\n')
    f.write(space*3+'<PmtMtd>TRF</PmtMtd>\n')
    f.write(space*3+'<PmtTpInf>\n')
    #level 4
    f.write(space*4+'<SvcLvl>\n')
    #level 5
    f.write(space*5+'<Cd>SDVA</Cd>\n')
    f.write(space*4+'</SvcLvl>\n')
    f.write(space*3+'</PmtTpInf>\n')
    #level 3
    f.write(space*3+'<ReqdExctnDt>'+next_friday.strftime("%Y-%m-%d")+'</ReqdExctnDt>\n')
    #DEBTOR ACCOUNT *****
    f.write(space*3+'<Dbtr>\n')
    #level 4
    f.write(space*4+'<Nm>Company Name</Nm>\n')
    f.write(space*4+'<PstlAdr>\n')
    #level 5
    f.write(space*5+'<StrtNm>Company Address</StrtNm>\n')
    f.write(space*5+'<PstCd>adr cont</PstCd>\n')
    f.write(space*5+'<TwnNm>city</TwnNm>\n')
    f.write(space*5+'<Ctry>ct</Ctry>\n')
    #end level 4
    f.write(space*4+'</PstlAdr>\n')
    #end level 3
    f.write(space*3+'</Dbtr>\n')
    #level 3
    f.write(space*3+'<DbtrAcct>\n')
    #level 4
    f.write(space*4+'<Id>\n')
    #level 5
    f.write(space*5+'<IBAN>IBAN INFO</IBAN>\n')
    #end level 4
    f.write(space*4+'</Id>\n')
    f.write(space*4+'<Ccy> ccy </Ccy>\n')
    #end level 3
    f.write(space*3+'</DbtrAcct>\n')
    #level 3
    f.write(space*3+'<DbtrAgt>\n')
    #level 4
    f.write(space*4+'<FinInstnId>\n')
    #level 5
    f.write(space*5+'<BIC>BIC code</BIC>\n')
    #end level 4
    f.write(space*4+'</FinInstnId>\n')
    #end level 3
    f.write(space*3+'</DbtrAgt>\n')
    #level 3
    f.write(space*3+'<ChrgBr>SHAR</ChrgBr>\n')
    #CUSTOMER TRANSACTIONS *****
    #level 3
    #for row in excel file  ****** the following will be indented under this for loop *****************
    count = 1
    for row in ws.iter_rows(min_row=2, max_col=13):
        row = [cell.value for cell in row]
        if row[11] == None:
            row[11] = " "
        if row[6] == None:
            row[6] = " "
        if row[1] == None:
            row[1] = " "
        if row[2] == None:
            row[2] = " "
        if row[10] == None:
            row[10] = " "
        if row[7] == None:
            row[7] = " "
        f.write(space*3+'<CdtTrfTxInf>\n')
        #level 4
        f.write(space*4+'<PmtId>\n')
        #level 5
        f.write(space*5+'<EndToEndId>'+ PaymentID +'</EndToEndId>\n')
        f.write(space*4+'</PmtId>\n')
        #level 4
        f.write(space*4+'<Amt>\n')
        #level 5
        f.write(space*5+'<InstdAmt Ccy="ccy">'+str(row[11])+'</InstdAmt>\n')
        f.write(space*4+'</Amt>\n')
        #level 4
        f.write(space*4+'<CdtrAgt>\n')
        #level 5
        f.write(space*5+'<FinInstnId>\n')
        #level 6
        f.write(space*6+'<BIC>'+str(row[6])+'</BIC>\n')
        f.write(space*5+'</FinInstnId>\n')
        f.write(space*4+'</CdtrAgt>\n')
        #level 4
        f.write(space*4+'<Cdtr>\n')
        #level 5
        nameStr = str(row[1]) + " " + str(row[2])
        nameStr2 = nameStr.rstrip()
        f.write(space*5+'<Nm>'+nameStr2+'</Nm>\n')
        f.write(space*5+'<PstlAdr>\n')
        #level 6
        f.write(space*6+'<Ctry>'+str(row[10])+'</Ctry>\n')
        f.write(space*5+'</PstlAdr>\n')
        f.write(space*4+'</Cdtr>\n')
        #level 4
        f.write(space*4+'<CdtrAcct>\n')
        #level 5
        f.write(space*5+'<Id>\n')
        #level 6
        f.write(space*6+'<IBAN>'+str(row[7])+'</IBAN>\n')
        f.write(space*5+'</Id>\n')
        f.write(space*4+'</CdtrAcct>\n')
        #level 4
        f.write(space*4+'<RmtInf>\n')
        #level 5
        f.write(space*5+'<Ustrd>'+'Payments' + next_friday.strftime("%#m-%#d-%Y")+'</Ustrd>\n')
        f.write(space*4+'</RmtInf>\n')
        #end customer transactions
        f.write(space*3+'</CdtTrfTxInf>\n')
    
    
    #END PAYMENT INFO
    #end level 2
    f.write(space*2+'</PmtInf>\n')
    #end level 1
    f.write(space+'</CstmrCdtTrfInitn>\n')
    #end top level (END DOCUMENT)
    f.write('</Document>')


    f.close

main()
